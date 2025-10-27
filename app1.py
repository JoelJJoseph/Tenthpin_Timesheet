!pip install openpyxl
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import io

def check_timesheet(file):
    """
    Process the uploaded Excel file and return the modified workbook
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file)
    ws = wb['Sheet1']
    
    # Clear previous formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill(fill_type=None)
    
    # Define header row
    header_row = 5
    
    # Find last row and column
    last_row = ws.max_row
    last_col = ws.max_column
    
    billable_hours_total = 0
    non_billable_hours_total = 0
    
    not_filled = []
    non_billable = []
    
    # First pass: Identify columns to exclude
    empty_columns = []
    valid_date_columns = []
    
    for col in range(2, last_col + 1):
        header_val = str(ws.cell(header_row, col).value or "").strip()
        
        # Check if column is blank or Grand Total
        is_blank_column = "(blank)" in header_val.lower() or "grand total" in header_val.lower()
        
        if is_blank_column:
            empty_columns.append(col)
        elif ws.cell(header_row, col).value and isinstance(ws.cell(header_row, col).value, datetime):
            # This is a valid date column
            valid_date_columns.append(col)
            
            # Check if column is completely empty
            is_column_empty = True
            for row in range(6, last_row + 1):
                employee = ws.cell(row, 1).value
                
                if employee and "total" not in str(employee).lower():
                    cell_val = ws.cell(row, col).value
                    
                    if cell_val and str(cell_val).strip() and cell_val != 0:
                        is_column_empty = False
                        break
            
            if is_column_empty:
                empty_columns.append(col)
    
    # Second pass: Check each employee row
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    
    for row in range(6, last_row + 1):
        employee = ws.cell(row, 1).value
        
        # Skip empty rows or totals
        if not employee or "total" in str(employee).lower():
            continue
        
        row_has_issue = False
        emp_billable = 0
        emp_non_billable = 0
        missing_dates_for_employee = []
        
        # Loop through valid date columns
        for col in valid_date_columns:
            # Skip excluded columns
            if col in empty_columns:
                continue
            
            date_val = ws.cell(header_row, col).value
            cell_val = ws.cell(row, col).value
            
            # Check if cell is empty or zero
            if not cell_val or str(cell_val).strip() == "" or cell_val == 0:
                missing_dates_for_employee.append(date_val)
                row_has_issue = True
                
            elif isinstance(cell_val, datetime):
                # Calculate hours
                hours = cell_val.hour + cell_val.minute / 60 + cell_val.second / 3600
                
                if hours == 0:
                    non_billable.append([employee, date_val])
                    emp_non_billable += 0
                else:
                    emp_billable += hours
            else:
                non_billable.append([employee, date_val])
        
        # Add to not_filled collection with all missing dates
        if missing_dates_for_employee:
            dates_list = ", ".join([d.strftime("%m/%d/%Y") for d in missing_dates_for_employee])
            not_filled.append([employee, dates_list])
        
        # Highlight entire row if issues found
        if row_has_issue:
            for col in range(1, last_col + 1):
                ws.cell(row, col).fill = red_fill
        
        billable_hours_total += emp_billable
        non_billable_hours_total += emp_non_billable
    
    # Create Summary sheet
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    
    summary_ws = wb.create_sheet("Summary")
    
    # Format summary sheet
    r = 1
    
    summary_ws.cell(r, 1).value = "Employees who didn't fill timesheet"
    summary_ws.cell(r, 1).font = openpyxl.styles.Font(bold=True, size=12)
    r += 1
    
    summary_ws.cell(r, 1).value = "Employee Name"
    summary_ws.cell(r, 2).value = "Missing Dates"
    summary_ws.cell(r, 1).font = openpyxl.styles.Font(bold=True)
    summary_ws.cell(r, 2).font = openpyxl.styles.Font(bold=True)
    r += 1
    
    for item in not_filled:
        summary_ws.cell(r, 1).value = item[0]
        summary_ws.cell(r, 2).value = item[1]
        r += 1
    
    r += 1
    summary_ws.cell(r, 1).value = "Employees who logged non-billable hours"
    summary_ws.cell(r, 1).font = openpyxl.styles.Font(bold=True)
    r += 1
    
    summary_ws.cell(r, 1).value = "Employee Name"
    summary_ws.cell(r, 2).value = "Date"
    summary_ws.cell(r, 1).font = openpyxl.styles.Font(bold=True)
    summary_ws.cell(r, 2).font = openpyxl.styles.Font(bold=True)
    r += 1
    
    for item in non_billable:
        summary_ws.cell(r, 1).value = item[0]
        summary_ws.cell(r, 2).value = item[1]
        summary_ws.cell(r, 2).number_format = 'M/D/YYYY'
        r += 1
    
    r += 1
    summary_ws.cell(r, 1).value = "Total Billable Hours"
    summary_ws.cell(r, 1).font = openpyxl.styles.Font(bold=True)
    summary_ws.cell(r, 2).value = billable_hours_total
    r += 1
    
    summary_ws.cell(r, 1).value = "Total Non-Billable Hours"
    summary_ws.cell(r, 1).font = openpyxl.styles.Font(bold=True)
    summary_ws.cell(r, 2).value = non_billable_hours_total
    
    # Auto-fit columns
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb, len(not_filled), billable_hours_total, non_billable_hours_total


# Streamlit App
def main():
    st.set_page_config(page_title="Timesheet Checker", page_icon="📊", layout="wide")
    
    st.title("📊 Timesheet Validation Tool")
    st.markdown("---")
    
    st.markdown("""
    ### How to use:
    1. Upload your timesheet Excel file
    2. Click 'Process Timesheet'
    3. Review the results and download the processed file
    
    **Note:** The tool will:
    - Highlight rows in red where employees haven't filled all 5 working days
    - Create a summary sheet with missing entries
    - Calculate billable and non-billable hours
    """)
    
    st.markdown("---")
    
    # File upload
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        st.success(f"✅ File uploaded: {uploaded_file.name}")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if st.button("🔍 Process Timesheet", type="primary", use_container_width=True):
                with st.spinner("Processing timesheet..."):
                    try:
                        # Process the file
                        wb, missing_count, billable_hours, non_billable_hours = check_timesheet(uploaded_file)
                        
                        # Save to BytesIO
                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)
                        
                        # Display results
                        st.markdown("---")
                        st.subheader("📈 Processing Results")
                        
                        # Metrics
                        metric_col1, metric_col2, metric_col3 = st.columns(3)
                        
                        with metric_col1:
                            st.metric("Employees with Missing Entries", missing_count)
                        
                        with metric_col2:
                            st.metric("Total Billable Hours", f"{billable_hours:.2f}")
                        
                        with metric_col3:
                            st.metric("Total Non-Billable Hours", f"{non_billable_hours:.2f}")
                        
                        st.markdown("---")
                        
                        # Download button
                        st.subheader("📥 Download Processed File")
                        
                        download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
                        
                        with download_col2:
                            st.download_button(
                                label="⬇️ Download Processed Excel File",
                                data=output,
                                file_name=f"processed_timesheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True
                            )
                        
                        st.success("✅ Processing completed successfully!")
                        
                        st.info("💡 The processed file contains:\n"
                                "- Highlighted rows (in red) for incomplete timesheets\n"
                                "- A new 'Summary' sheet with detailed reports")
                        
                    except Exception as e:
                        st.error(f"❌ Error processing file: {str(e)}")
                        st.exception(e)
    
    else:
        st.info("👆 Please upload an Excel file to begin")
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: gray;'>
    <small>Timesheet Validation Tool | Built with Streamlit</small>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":

    main()
